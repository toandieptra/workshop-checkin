import csv
import io
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..db import get_db
from ..models import Guest, Workshop
from ..auth.dependencies import require_permission

router = APIRouter(prefix="/api", tags=["import-export"])

COLS = ["full_name", "phone", "email", "business_model", "role_title", "guest_type"]


def _parse_int(v, default: int = 1) -> int:
    try:
        n = int(float(str(v).strip()))
        return n if n > 0 else default
    except (TypeError, ValueError):
        return default


@router.post("/workshops/{workshop_id}/import", dependencies=[Depends(require_permission("guests.write"))])
async def import_guests(workshop_id: uuid.UUID, file: UploadFile = File(...), db: AsyncSession = Depends(get_db)):
    from ..services.registration_confirmation import apply_registration_policy
    from ..services.zbs import normalize_phone
    w = await db.get(Workshop, workshop_id)
    if not w:
        raise HTTPException(404, "workshop not found")
    data = await file.read()
    name = (file.filename or "").lower()

    rows: list[dict] = []
    if name.endswith(".xlsx") or name.endswith(".xls"):
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for r in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})
    else:
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]

    created = 0
    confirmed = 0
    for row in rows:
        full_name = (row.get("full_name") or row.get("ten") or "").strip()
        if not full_name:
            continue
        g = Guest(
            workshop_id=workshop_id,
            full_name=full_name,
            phone=normalize_phone(row.get("phone")) or None,
            email=(row.get("email") or None),
            business_model=(row.get("business_model") or row.get("mô hình kinh doanh") or row.get("mo hinh kinh doanh") or None),
            registered_at=datetime.now(timezone.utc),
            role_title=(row.get("role_title") or None),
            guest_type=(row.get("guest_type") or None),
            party_size=_parse_int(
                row.get("party_size") or row.get("so_khach") or row.get("số khách")
                or row.get("số vé đăng ký") or row.get("so ve") or 1
            ),
            registration_status="pending",
        )
        db.add(g)
        await db.flush()
        if await apply_registration_policy(db, g, auto_confirm=w.auto_confirm_registration):
            confirmed += 1
        created += 1
    await db.commit()
    return {
        "imported": created,
        "confirmed": confirmed,
        "pending": created - confirmed,
        "total_rows": len(rows),
    }


@router.get("/export/guests", dependencies=[Depends(require_permission("guests.export"))])
async def export_guests(
    db: AsyncSession = Depends(get_db),
    workshop_id: str | None = Query(default=None, description="uuid hoặc 'all'"),
    workshop_ids: str | None = Query(default=None, description="Danh sách UUID workshop, cách nhau bằng dấu phẩy"),
    status: str = Query(default="all", pattern="^(all|checked_in|not_checked_in)$"),
):
    """Xuất khách ra file .xlsx.

    - workshop_id: uuid của workshop, hoặc 'all' / None để xuất tất cả workshop
      (khi đó thêm cột 'workshop' để biết khách thuộc workshop nào).
    - status: 'all' | 'checked_in' | 'not_checked_in'.
    """
    stmt = (
        select(Guest, Workshop.name)
        .join(Workshop, Guest.workshop_id == Workshop.id)
        .where(Guest.deleted_at.is_(None))
    )
    if workshop_ids:
        try:
            workshop_uuid_list = [uuid.UUID(value.strip()) for value in workshop_ids.split(",") if value.strip()]
        except ValueError:
            raise HTTPException(400, "workshop_ids không hợp lệ")
        if not workshop_uuid_list:
            raise HTTPException(400, "workshop_ids không hợp lệ")
        stmt = stmt.where(Guest.workshop_id.in_(workshop_uuid_list))
    elif workshop_id and workshop_id != "all":
        try:
            wid_uuid = uuid.UUID(workshop_id)
        except ValueError:
            raise HTTPException(400, "workshop_id không hợp lệ")
        stmt = stmt.where(Guest.workshop_id == wid_uuid)
    if status == "checked_in":
        stmt = stmt.where(Guest.checkin_status == "checked_in")
    elif status == "not_checked_in":
        stmt = stmt.where(Guest.checkin_status != "checked_in")
    stmt = stmt.order_by(Guest.checkin_status.desc(), Guest.full_name)

    rows = (await db.execute(stmt)).all()

    def _fmt(dt):
        # openpyxl không hỗ trợ datetime có tzinfo — Postgres trả về aware
        # datetime khi cột khai báo DateTime(timezone=True). Convert sang
        # UTC naive để Excel chấp nhận, đồng thời giữ giá trị giờ chính xác.
        if dt is None:
            return None
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt

    def _checkin_label(s: str | None) -> str:
        return "Đã check-in" if s == "checked_in" else "Chưa check-in"

    def _sync_label(s: str | None) -> str:
        return {
            "synced": "Đã đồng bộ",
            "pending_push": "Chờ đồng bộ",
            "error": "Lỗi đồng bộ",
        }.get(s or "", s or "")

    headers = [
        "workshop", "full_name", "phone", "email", "company", "business_model",
        "role_title", "guest_type", "party_size", "note",
        "checkin_status", "checked_in_at",
        "sync_status", "lark_record_id",
        "registered_at", "created_at",
    ]
    body_rows = []
    for g, w_name in rows:
        body_rows.append([
            w_name,
            g.full_name,
            g.phone or "",
            g.email or "",
            g.company or "",
            g.business_model or "",
            g.role_title or "",
            g.guest_type or "",
            g.party_size,
            g.note or "",
            _checkin_label(g.checkin_status),
            _fmt(g.checked_in_at),
            _sync_label(g.sync_status),
            g.lark_record_id or "",
            _fmt(g.registered_at),
            _fmt(g.created_at),
        ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Guests"

    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF0D3B42")
    header_align = Alignment(horizontal="left", vertical="center")

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    date_fmt = "yyyy-mm-dd hh:mm:ss"
    date_col_indexes = [headers.index(h) + 1 for h in ("checked_in_at", "registered_at", "created_at")]
    for row in body_rows:
        ws.append(row)
        excel_row = ws.max_row
        for col_idx in date_col_indexes:
            ws.cell(row=excel_row, column=col_idx).number_format = date_fmt

    sample_for_width = body_rows[:200]
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in sample_for_width:
            val = row[col_idx - 1]
            if val is None:
                continue
            s = val.strftime("%Y-%m-%d %H:%M:%S") if hasattr(val, "strftime") else str(val)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"guests_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
