import asyncio
import io
import uuid
from datetime import datetime, timezone

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.models import Guest
from app.routers.import_export import export_guests


class Result:
    def __init__(self, rows):
        self.rows = rows

    def all(self):
        return self.rows


class FakeDb:
    def __init__(self, guest_rows, note_rows):
        self.results = iter((Result(guest_rows), Result(note_rows)))
        self.statements = []

    async def execute(self, stmt):
        self.statements.append(stmt)
        return next(self.results)


async def response_bytes(response):
    return b"".join([chunk async for chunk in response.body_iterator])


def test_export_guests_includes_structured_notes_and_legacy_fallback():
    structured_guest = Guest(
        id=uuid.uuid4(),
        workshop_id=uuid.uuid4(),
        full_name="Nguyễn Văn A",
        note="Ghi chú legacy không được ưu tiên",
        source="Đại lý giới thiệu",
        party_size=1,
        checkin_status="not_checked_in",
    )
    legacy_guest = Guest(
        id=uuid.uuid4(),
        workshop_id=uuid.uuid4(),
        full_name="Trần Thị B",
        note="Chỉ có ghi chú legacy",
        source="Khác",
        source_detail="Bạn bè giới thiệu",
        party_size=1,
        checkin_status="checked_in",
    )
    note_rows = [
        (structured_guest.id, "Ghi chú mới nhất", datetime(2026, 7, 28, 3, 4, 5, tzinfo=timezone.utc), "Admin"),
        (structured_guest.id, "Ghi chú cũ hơn", datetime(2026, 7, 27, 2, 3, 4, tzinfo=timezone.utc), None),
    ]
    db = FakeDb(
        [(structured_guest, "Workshop A"), (legacy_guest, "Workshop B")],
        note_rows,
    )

    response = asyncio.run(export_guests(
        db=db,
        workshop_id=None,
        workshop_ids=None,
        status="all",
        business_model=None,
        source=None,
    ))
    workbook = load_workbook(io.BytesIO(asyncio.run(response_bytes(response))))
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[1]]
    note_column = headers.index("note") + 1
    source_column = headers.index("source") + 1

    assert worksheet.cell(2, note_column).value == (
        "Admin · 10:04:05 28/7/2026: Ghi chú mới nhất\n"
        "— · 09:03:04 27/7/2026: Ghi chú cũ hơn"
    )
    assert worksheet.cell(2, note_column).alignment.wrap_text is True
    assert worksheet.cell(3, note_column).value == "Chỉ có ghi chú legacy"
    assert worksheet.cell(2, source_column).value == "Đại lý giới thiệu"
    assert worksheet.cell(3, source_column).value == "Khác: Bạn bè giới thiệu"


def test_export_guests_applies_business_model_and_source_filters():
    db = FakeDb([], [])

    asyncio.run(export_guests(
        db=db,
        workshop_id=None,
        workshop_ids=None,
        status="all",
        business_model="Khác",
        source="Đại lý giới thiệu",
    ))

    query = db.statements[0].compile()
    sql = str(query)
    assert "trim(coalesce(guests.business_model" in sql
    assert "NOT IN" in sql
    assert "guests.source =" in sql
    assert "Đại lý giới thiệu" in query.params.values()


@pytest.mark.parametrize(
    ("business_model", "source", "message"),
    [
        ("Không hợp lệ", None, "Mô hình kinh doanh không hợp lệ"),
        (None, "Không hợp lệ", "Nguồn khách không hợp lệ"),
    ],
)
def test_export_guests_rejects_invalid_filters(business_model, source, message):
    db = FakeDb([], [])

    with pytest.raises(HTTPException) as exc:
        asyncio.run(export_guests(
            db=db,
            workshop_id=None,
            workshop_ids=None,
            status="all",
            business_model=business_model,
            source=source,
        ))

    assert exc.value.status_code == 400
    assert exc.value.detail == message
    assert db.statements == []
